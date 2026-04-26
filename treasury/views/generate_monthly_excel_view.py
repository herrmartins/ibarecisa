import io
from datetime import datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404

from treasury.models import AccountingPeriod, TransactionModel


BRL_FORMAT = '#,##0.00_);[Red](#,##0.00)'
DATE_FORMAT = 'DD/MM/YYYY'

HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='2F5496')
SUBTITLE_FONT = Font(name='Calibri', bold=True, size=11, color='2F5496')
NORMAL_FONT = Font(name='Calibri', size=11)
BOLD_FONT = Font(name='Calibri', bold=True, size=11)
ENTRY_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
EXIT_FILL = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
TOTAL_FILL = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def _is_treasury_user(user):
    if not user.is_authenticated:
        return False
    if user.type == "REGULAR":
        return True
    return (
        user.is_treasurer
        or user.is_secretary
        or user.is_pastor
        or user.is_staff
        or user.is_superuser
    )


def _apply_cell_style(cell, font=None, fill=None, alignment=None, number_format=None, border=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if number_format:
        cell.number_format = number_format
    if border:
        cell.border = border


@login_required
def generate_monthly_excel(request, year, month):
    if not _is_treasury_user(request.user):
        return HttpResponseForbidden("Sem permissao.")

    period = get_object_or_404(
        AccountingPeriod,
        month__year=year,
        month__month=month,
    )

    transactions = TransactionModel.objects.filter(
        accounting_period=period,
        transaction_type='original',
    ).select_related('category').order_by('date', 'created_at')

    summary = period.get_transactions_summary()
    opening_balance = period.opening_balance or Decimal('0.00')
    total_positive = summary['total_positive']
    total_negative = summary['total_negative']
    net = summary['net']
    closing_balance = period.closing_balance if period.closing_balance else (opening_balance + net)

    wb = Workbook()

    _build_summary_sheet(wb.active, period, opening_balance, total_positive, total_negative, net, closing_balance, summary['count'])
    _build_transactions_sheet(wb.create_sheet('Transacoes'), period, transactions, opening_balance)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    month_str = period.month.strftime('%Y_%m')
    filename = f"relatorio_mensal_{month_str}.xlsx"

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _build_summary_sheet(ws, period, opening_balance, total_positive, total_negative, net, closing_balance, count):
    ws.title = 'Resumo'

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25

    row = 1
    ws.merge_cells('A1:B1')
    title_cell = ws.cell(row=row, column=1, value=f'Relatorio Mensal - {period.month_name}/{period.year}')
    _apply_cell_style(title_cell, font=TITLE_FONT, alignment=Alignment(horizontal='center'))
    row += 2

    labels_values = [
        ('Mes/Ano', f'{period.month_name}/{period.year}'),
        ('Status', period.get_status_display()),
        ('Saldo Anterior (Abertura)', float(opening_balance)),
        ('Total de Entradas', float(total_positive)),
        ('Total de Saida', float(total_negative)),
        ('Resultado do Mes', float(net)),
        ('Saldo Final', float(closing_balance)),
        ('Quantidade de Transacoes', count),
    ]

    for label, value in labels_values:
        label_cell = ws.cell(row=row, column=1, value=label)
        value_cell = ws.cell(row=row, column=2, value=value)

        _apply_cell_style(label_cell, font=BOLD_FONT, border=THIN_BORDER, alignment=Alignment(horizontal='right', vertical='center'))

        is_currency = isinstance(value, float)
        _apply_cell_style(
            value_cell,
            font=BOLD_FONT if label in ('Resultado do Mes', 'Saldo Final') else NORMAL_FONT,
            border=THIN_BORDER,
            alignment=Alignment(horizontal='right', vertical='center'),
            number_format=BRL_FORMAT if is_currency else None,
        )

        if label in ('Resultado do Mes', 'Saldo Final'):
            _apply_cell_style(label_cell, fill=TOTAL_FILL)
            _apply_cell_style(value_cell, fill=TOTAL_FILL)

        row += 1

    row += 1
    ws.cell(row=row, column=1, value='Gerado em:').font = Font(name='Calibri', size=9, italic=True, color='666666')
    ws.cell(row=row, column=2, value=datetime.now().strftime('%d/%m/%Y %H:%M')).font = Font(name='Calibri', size=9, italic=True, color='666666')


def _build_transactions_sheet(ws, period, transactions, opening_balance):
    ws.title = 'Transacoes'

    columns = [
        ('Data', 14),
        ('Descricao', 40),
        ('Categoria', 20),
        ('Tipo', 12),
        ('Valor', 18),
        ('Saldo Acumulado', 18),
    ]

    for col_idx, (header, width) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        cell = ws.cell(row=1, column=col_idx, value=header)
        _apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL, alignment=Alignment(horizontal='center', vertical='center'), border=THIN_BORDER)

    ws.auto_filter.ref = f'A1:{get_column_letter(len(columns))}1'
    ws.freeze_panes = 'A2'

    running_balance = float(opening_balance)
    row = 2

    for tx in transactions:
        amount = float(tx.amount)
        if tx.is_positive:
            tx_type = 'Entrada'
            running_balance += amount
            row_fill = ENTRY_FILL
        else:
            tx_type = 'Saida'
            running_balance -= amount
            row_fill = EXIT_FILL

        values = [
            tx.date,
            tx.description,
            tx.category.name if tx.category else 'Sem categoria',
            tx_type,
            amount if tx.is_positive else -amount,
            running_balance,
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            _apply_cell_style(cell, font=NORMAL_FONT, border=THIN_BORDER, fill=row_fill)

            if col_idx == 1:
                cell.number_format = DATE_FORMAT
                cell.alignment = Alignment(horizontal='center')
            elif col_idx in (5, 6):
                cell.number_format = BRL_FORMAT
                cell.alignment = Alignment(horizontal='right')
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal='center')

        row += 1

    row += 1

    total_label_cell = ws.cell(row=row, column=1, value='TOTAIS')
    _apply_cell_style(total_label_cell, font=BOLD_FONT, fill=TOTAL_FILL, border=THIN_BORDER)
    for col in range(2, 5):
        c = ws.cell(row=row, column=col)
        _apply_cell_style(c, fill=TOTAL_FILL, border=THIN_BORDER)

    tx_count = transactions.count()
    pos_sum = sum(float(t.amount) for t in transactions if t.is_positive)
    neg_sum = sum(float(t.amount) for t in transactions if not t.is_positive)

    ws.cell(row=row, column=5, value=pos_sum - neg_sum).number_format = BRL_FORMAT
    _apply_cell_style(ws.cell(row=row, column=5), font=BOLD_FONT, fill=TOTAL_FILL, border=THIN_BORDER, alignment=Alignment(horizontal='right'))

    ws.cell(row=row, column=6, value=running_balance).number_format = BRL_FORMAT
    _apply_cell_style(ws.cell(row=row, column=6), font=BOLD_FONT, fill=TOTAL_FILL, border=THIN_BORDER, alignment=Alignment(horizontal='right'))

    row += 1
    info_row = row
    ws.cell(row=info_row, column=1, value=f'{tx_count} transacao(oes)').font = Font(name='Calibri', size=9, italic=True, color='666666')
    ws.cell(row=info_row, column=5, value=f'Entradas: {pos_sum:,.2f}').font = Font(name='Calibri', size=9, italic=True, color='548235')
    ws.cell(row=info_row, column=5).number_format = BRL_FORMAT
    ws.cell(row=info_row + 1, column=5, value=f'Saidas: {neg_sum:,.2f}').font = Font(name='Calibri', size=9, italic=True, color='C00000')
    ws.cell(row=info_row + 1, column=5).number_format = BRL_FORMAT
